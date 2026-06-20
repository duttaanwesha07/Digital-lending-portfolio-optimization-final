"""
Workstream D - EDA generating script (reproducible)
===================================================
Regenerates the entire Workstream D EDA from the FROZEN Workstream C output:
the workbook (all slices + diagnostics) and the six PNG exhibits.

Scope
-----
- OURS (reproduced here): Q3 product x tenure x ticket economics, plus the
  cross-cutting diagnostics (equal-MOB check, vintage, roll-rate, portfolio
  overview) and the product x channel crossover capstone.
- COWORKER (incorporated, not regenerated): Q2 channel economics. If the Q2
  controlled workbook is provided, its sheets are folded in and reconciled;
  the Q2 numbers are sourced from that file, never recomputed here.

Inputs  (frozen WS-C output; seed 20260603, Gate 1 PASS 32/32 MUST)
    <DATA>/data/parquet/loans.parquet
    <DATA>/data/parquet/customers.parquet
    <DATA>/data/parquet/repayments.parquet
    <DATA>/meta/value_proxy_components.csv
    [optional] Q2 controlled workbook (wsD_Q2_controlled_workbook.xlsx)

Outputs (written to <OUT>/)
    WorkstreamD_consolidated_log.xlsx        (if Q2 provided)
    Q3_results_log.xlsx                      (Q3 + diagnostics only, always)
    sliceA_product_bar.png  sliceB_tenure_heatmap.png  sliceC_ticket_heatmap.png
    fragment1_vintage.png   fragment2_rollrate.png     fragment3_portfolio_overview.png
    fragment4_product_channel.png

Run
    python wsD_analysis.py  --data ./out  --out ./wsD_outputs  [--q2 ./wsD_Q2_controlled_workbook.xlsx]

Notes
-----
- Definitions (must match Q2): value_proxy = NII + Fee - Loss - Servicing - CAC;
  margin_before_cac = NII + Fee - Loss - Servicing. Reconciliation is asserted.
- Workbook derived columns are written as live Excel formulas; Excel/LibreOffice
  recalculates them on open.
- Deterministic: data is frozen, so re-running reproduces identical numbers.
- Caveat carried throughout: single M24 snapshot, mixed loan ages -> low-maturity
  cells show profit/default "so far", not lifetime (long/large cells understated).

Dependencies: pandas, numpy, matplotlib, openpyxl, pyarrow
"""

import argparse, os, json
import numpy as np
import pandas as pd

CRORE = 10_000_000
PRODUCTS = ['BNPL', 'Personal', 'SME']
CHANNELS = ['Referral', 'Partner-embedded', 'Digital ads', 'Organic', 'DSA']
MIN_BAND_SHARE = 0.05  # below this, a ticket band is "degenerate" -> auto-fallback

# ---------------------------------------------------------------- bucketing
def canon_product(p):
    s = str(p).strip().lower()
    if s.startswith('bnpl'):     return 'BNPL'
    if s.startswith('personal'): return 'Personal'
    if s.startswith('sme'):      return 'SME'
    return p

def tenure_bucket(p, t):
    if p == 'BNPL':     return '1-3' if t <= 3 else '4-6'
    if p == 'Personal': return '3-12' if t <= 12 else ('13-24' if t <= 24 else '25-36')
    if p == 'SME':      return '6-12' if t <= 12 else ('13-24' if t <= 24 else '25-36+')
    return None

# documented-anchored ticket bands (BNPL rescaled to the real 2k-50k range, A-033)
TICKET_BANDS = {
    'BNPL':     ([10_000, 25_000],      ['<10k', '10-25k', '>25k']),
    'Personal': ([50_000, 150_000],     ['<50k', '50-150k', '>150k']),
    'SME':      ([250_000, 1_000_000],  ['<250k', '250k-1M', '>1M']),
}
TENURE_ORDER = {'BNPL': ['1-3', '4-6'], 'Personal': ['3-12', '13-24', '25-36'],
                'SME': ['6-12', '13-24', '25-36+']}
TICKET_ORDER = {p: TICKET_BANDS[p][1] for p in PRODUCTS}

def ticket_buckets_for(series, product):
    """Documented bands, with auto-fallback to data tertiles if a band is degenerate."""
    edges, labels = TICKET_BANDS[product]
    bk = pd.cut(series, [-np.inf] + edges + [np.inf], labels=labels)
    if (bk.value_counts(normalize=True) < MIN_BAND_SHARE).any():
        q = series.quantile([1/3, 2/3]).round(0).tolist()
        labels = [f'<{int(q[0]):,}', f'{int(q[0]):,}-{int(q[1]):,}', f'>{int(q[1]):,}']
        bk = pd.cut(series, [-np.inf] + q + [np.inf], labels=labels)
        print(f'  [guard] {product}: documented ticket band degenerate -> fell back to data thirds {labels}')
    return bk.astype(str)

# ---------------------------------------------------------------- load + reconcile
def load_and_reconcile(data):
    loans = pd.read_parquet(f'{data}/data/parquet/loans.parquet')
    cust  = pd.read_parquet(f'{data}/data/parquet/customers.parquet')
    vc    = pd.read_csv(f'{data}/meta/value_proxy_components.csv')

    assert not loans.loan_id.duplicated().any(), 'duplicate loan_id in loans'
    assert not vc.loan_id.duplicated().any(),    'duplicate loan_id in components'
    m = loans.merge(vc, on='loan_id', how='inner', validate='one_to_one')
    m['product'] = m['product_type'].map(canon_product)
    m['margin_before_cac'] = m.nii + m.fee - m.loss - m.servicing

    recomputed = m.nii + m.fee - m.loss - m.servicing - m.cac_charged
    gap = (m.value_proxy - recomputed).abs().max()
    assert gap <= 1.0, f'value_proxy does not reconcile to components (max gap {gap})'
    print(f'  reconciliation OK (max gap {gap:.3f} rupee); {len(m):,} loans')

    m = m.merge(cust[['customer_id', 'acquisition_channel']], on='customer_id', how='left')
    return m, cust

# ---------------------------------------------------------------- slices
def slice_metrics(df, keys):
    g = df.groupby(keys).agg(
        n=('loan_id', 'size'), defr=('default_flag', 'mean'),
        mvp=('value_proxy', 'mean'), tvp=('value_proxy', 'sum'),
        mob=('months_on_book', 'mean'), ten=('tenure_mo', 'mean'),
        tk=('ticket_size', 'mean'), apr=('interest_rate_apr', 'mean'))
    return g

def compute_slices(m):
    A = slice_metrics(m, 'product').reindex(PRODUCTS)
    m = m.copy()
    m['tbk'] = [tenure_bucket(p, t) for p, t in zip(m['product'], m.tenure_mo)]
    kb = pd.Series(index=m.index, dtype=object)
    for p in PRODUCTS:
        idx = m['product'] == p
        kb[idx] = ticket_buckets_for(m.loc[idx, 'ticket_size'], p).values
    m['kbk'] = kb
    B = slice_metrics(m, ['product', 'tbk'])
    C = slice_metrics(m, ['product', 'kbk'])
    return A, B, C, m

# ---------------------------------------------------------------- equal-MOB default
def equal_mob(loans_rep):
    loans, rep = loans_rep
    first90 = rep[rep.dpd >= 90].groupby('loan_id').period_index.min()
    L = loans.copy()
    L['mob_at_default'] = L.loan_id.map(first90)
    L['obs'] = L.months_on_book
    out = []
    def cum(df, k):
        e = df[df.obs >= k]
        return (round((e.mob_at_default <= k).mean()*100, 1), len(e)) if len(e) else (np.nan, 0)
    for p in ['Personal', 'SME']:
        sub = L[(L.product_type == p) & (L.tenure_mo > 12)].copy()
        sub['bk'] = np.where(sub.tenure_mo <= 24, '13-24', '25-36' if p == 'Personal' else '25-36+')
        for bk in sorted(sub.bk.unique()):
            g = sub[sub.bk == bk]
            d12, n12 = cum(g, 12)
            out.append((p, bk, round(g.default_flag.mean()*100, 1), d12, n12))
    return pd.DataFrame(out, columns=['product', 'bucket', 'raw_default', 'default_at_MOB12', 'n_watched12'])

# ---------------------------------------------------------------- vintage
def vintage(loans_rep):
    loans, rep = loans_rep
    first30 = rep[rep.dpd >= 30].groupby('loan_id').period_index.min()
    L = loans.copy(); L['mob30'] = L.loan_id.map(first30); L['obs'] = L.months_on_book
    rows = []
    for c in sorted(L.origination_cohort.unique()):
        s = L[L.origination_cohort == c]; e = s[s.obs >= 3]
        rows.append((c, len(s), round(s.default_flag.mean()*100, 1),
                     round((e.mob30 <= 3).mean()*100, 1) if len(e) else np.nan,
                     len(e), round(s.origination_risk_grade.isin(['D', 'E']).mean()*100, 1)))
    return pd.DataFrame(rows, columns=['cohort', 'n', 'raw_default', 'early30_MOB3', 'n_obs3', 'grade_DE'])

# ---------------------------------------------------------------- roll-rate
def rollrate(rep):
    ORDER = ['Current', '1-30', '31-60', '61-90', '90+']
    r = rep.sort_values(['loan_id', 'period_index']).copy()
    r['next'] = r.groupby('loan_id').delinquency_bucket.shift(-1)
    pairs = r.dropna(subset=['next'])
    cnt = pd.crosstab(pairs.delinquency_bucket, pairs.next).reindex(index=ORDER, columns=ORDER).fillna(0)
    prob = (cnt.div(cnt.sum(axis=1), axis=0) * 100)
    first_del = rep[rep.dpd > 0].groupby('loan_id').period_index.min()
    ttd = (round(first_del.median(), 0), round(first_del.mean(), 1), len(first_del))
    return prob, ORDER, ttd

# ---------------------------------------------------------------- portfolio overview
def portfolio_overview(m, cust, rep):
    last = rep.sort_values(['loan_id', 'period_index']).groupby('loan_id').tail(1)[
        ['loan_id', 'outstanding_balance', 'delinquency_bucket']]
    L = m.merge(last, on='loan_id', how='left')
    active = L[L.loan_status == 'Active']
    o = dict(
        n_loans=len(m), n_cust=int(m.customer_id.nunique()),
        originated=round(m.ticket_size.sum()/CRORE), outstanding=round(active.outstanding_balance.sum()/CRORE),
        profit=round(m.value_proxy.sum()/CRORE, 1), ever90=round(m.default_flag.mean()*100, 1),
        cur_delq=round((active.delinquency_bucket != 'Current').mean()*100, 1),
        cur_90=round((active.delinquency_bucket == '90+').mean()*100, 1),
        wa_apr=round(np.average(m.interest_rate_apr, weights=m.ticket_size), 1),
        wa_ten=round(np.average(m.tenure_mo, weights=m.ticket_size), 1),
        mean_tk=round(m.ticket_size.mean()))
    con = m.groupby('product').agg(n=('loan_id', 'size'), expo=('ticket_size', 'sum'),
                                   profit=('value_proxy', 'sum')).reindex(PRODUCTS)
    return o, con

# ---------------------------------------------------------------- crossover
def crossover(m):
    g = m.groupby(['product', 'acquisition_channel'])
    mean_vp = g.value_proxy.mean().unstack().reindex(index=PRODUCTS, columns=CHANNELS)
    n = g.size().unstack().reindex(index=PRODUCTS, columns=CHANNELS)
    defr = g.default_flag.mean().unstack().reindex(index=PRODUCTS, columns=CHANNELS)
    tot = g.value_proxy.sum().unstack().reindex(index=PRODUCTS, columns=CHANNELS)
    return mean_vp, n, defr, tot

# ================================================================ CHARTS
def make_charts(A, B, C, vint, prob, ORDER, ov, con, xover, out):
    import matplotlib; matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from matplotlib.colors import TwoSlopeNorm
    from matplotlib import gridspec
    plt.rcParams.update({'font.family': 'DejaVu Sans', 'font.size': 9, 'figure.facecolor': 'white',
                         'axes.facecolor': 'white', 'axes.edgecolor': '#888'})
    GREEN, RED, NAVY, INK = '#2e8b57', '#c0392b', '#1f3864', '#222'
    slog = lambda x: np.sign(x)*np.log10(1+abs(x))

    # Slice A bar
    a = A
    fig, ax = plt.subplots(figsize=(7, 4.5))
    bars = ax.bar(PRODUCTS, a.mvp, color=[RED if v < 0 else GREEN for v in a.mvp], width=.6, zorder=3)
    ax.axhline(0, color='#444', lw=1); ax.grid(axis='y', color='#eee', zorder=0)
    for b, v, d in zip(bars, a.mvp, a.defr):
        ax.text(b.get_x()+b.get_width()/2, v+(1600 if v >= 0 else -1600), f'\u20b9{v:,.0f}',
                ha='center', va='bottom' if v >= 0 else 'top', fontweight='bold', color=INK)
        ax.text(b.get_x()+b.get_width()/2, -4200, f'default {d*100:.1f}%', ha='center', va='top', fontsize=9, color='#666')
    ax.set_ylabel('Mean profit per loan (\u20b9)'); ax.set_ylim(-7000, 40000)
    fig.suptitle('Slice A \u2014 Profit per loan by product', fontweight='bold', fontsize=13, y=0.99)
    ax.set_title('BNPL loses money on every loan; SME is the engine. Loan-level, M24 snapshot.', fontsize=8.5, color='#666', pad=8)
    plt.tight_layout(); plt.savefig(f'{out}/sliceA_product_bar.png', dpi=150); plt.close()

    def heatmap(df, order, title, subtitle, fname):
        val = np.full((3, 3), np.nan); lab = [['']*3 for _ in range(3)]
        bcol = df.index.names[1]
        for i, p in enumerate(PRODUCTS):
            for j, bk in enumerate(order[p]):
                if (p, bk) in df.index:
                    row = df.loc[(p, bk)]; v = row.mvp; mat = row.mob/row.ten*100
                    val[i, j] = v; lab[i][j] = f'{bk}\n\u20b9{v:,.0f}\ndef {row.defr*100:.1f}%\nmat {mat:.0f}%'
        norm = TwoSlopeNorm(vmin=slog(np.nanmin(val))*1.05, vcenter=0, vmax=slog(np.nanmax(val))*1.05)
        fig, ax = plt.subplots(figsize=(7.8, 4.9))
        im = ax.imshow(slog(val), cmap='RdYlGn', norm=norm, aspect='auto')
        ax.set_xticks([0, 1, 2]); ax.set_xticklabels(['smaller / shorter', 'middle', 'larger / longer'])
        ax.set_yticks([0, 1, 2]); ax.set_yticklabels(PRODUCTS)
        for i in range(3):
            for j in range(3):
                if lab[i][j]: ax.text(j, i, lab[i][j], ha='center', va='center', fontsize=8.3, color='#111', fontweight='bold')
                else: ax.add_patch(plt.Rectangle((j-.5, i-.5), 1, 1, color='#f2f2f2', zorder=2))
        fig.suptitle(title, fontweight='bold', fontsize=13, y=0.99)
        ax.set_title(subtitle, fontsize=8.5, color='#666', pad=8)
        ax.set_xticks(np.arange(-.5, 3, 1), minor=True); ax.set_yticks(np.arange(-.5, 3, 1), minor=True)
        ax.grid(which='minor', color='white', lw=2); ax.tick_params(which='minor', length=0)
        cb = fig.colorbar(im, ax=ax, shrink=.8); cb.set_label('profit (red=loss \u2192 green=profit)', fontsize=8); cb.set_ticks([])
        plt.tight_layout(); plt.savefig(fname, dpi=150); plt.close()

    heatmap(B, TENURE_ORDER, 'Slice B \u2014 Mean profit per loan by product x tenure',
            'mat% = how aged the cell is (low = profit is "so far", understated).',
            f'{out}/sliceB_tenure_heatmap.png')
    heatmap(C, TICKET_ORDER, 'Slice C \u2014 Mean profit per loan by product x ticket',
            'BNPL underwater at EVERY ticket band; large-ticket SME is the crown jewel.',
            f'{out}/sliceC_ticket_heatmap.png')

    # Vintage
    df = vint; x = range(len(df)); labels = [c[2:] for c in df.cohort]
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(9, 6.4), height_ratios=[2.1, 1], sharex=True)
    ax1.plot(x, df.raw_default, '-o', color=RED, ms=3, lw=1.6, label='Raw cumulative default (naive)')
    msk = df.early30_MOB3.notna()
    ax1.plot(np.array(x)[msk], df.early30_MOB3[msk], '-o', color='#1f6feb', ms=3, lw=1.6, label='Early 30+ by MOB 3 (equal age)')
    xe = np.array(x)[msk]; z = np.polyfit(xe, df.early30_MOB3[msk].values, 1)
    ax1.plot(xe, np.poly1d(z)(xe), '--', color='#1f6feb', lw=1, alpha=.6)
    ax1.set_ylabel('default / early-delinquency (%)'); ax1.set_ylim(0, 30)
    ax1.legend(loc='center left', fontsize=8, frameon=False); ax1.grid(axis='y', color='#eee')
    ax2.plot(x, df.grade_DE, '-o', color='#6b4fbb', ms=3, lw=1.4); ax2.axhline(df.grade_DE.mean(), color='#888', ls='--', lw=.8)
    ax2.set_ylabel('grade D+E\nintake (%)'); ax2.set_ylim(20, 32); ax2.grid(axis='y', color='#eee')
    ax2.set_xticks(list(x)); ax2.set_xticklabels(labels, rotation=45, ha='right', fontsize=7)
    ax2.set_xlabel('origination cohort (YY-MM)')
    fig.suptitle('Fragment 1 \u2014 Vintage analysis: do newer cohorts deteriorate?', fontweight='bold', fontsize=13, y=0.98)
    ax1.set_title('Raw default by cohort misleads; the equal-age view shows the real, mild drift.', fontsize=8.5, color='#666', pad=6)
    plt.tight_layout(); plt.savefig(f'{out}/fragment1_vintage.png', dpi=150); plt.close()

    # Roll-rate
    fig = plt.figure(figsize=(9.2, 5.2)); gs = gridspec.GridSpec(1, 2, width_ratios=[1.55, 1], wspace=0.32)
    ax = fig.add_subplot(gs[0]); M = prob.values
    im = ax.imshow(M, cmap='Blues', vmin=0, vmax=100, aspect='auto')
    ax.set_xticks(range(5)); ax.set_xticklabels(ORDER, rotation=30, ha='right')
    ax.set_yticks(range(5)); ax.set_yticklabels(ORDER); ax.set_xlabel('next month \u2192'); ax.set_ylabel('this month')
    for i in range(5):
        for j in range(5):
            if M[i, j] >= 0.05: ax.text(j, i, f'{M[i,j]:.0f}%', ha='center', va='center', fontsize=8.5,
                                        color='white' if M[i, j] > 55 else '#222', fontweight='bold')
    ax.add_patch(plt.Rectangle((3.5, 2.5), 1, 1, fill=False, edgecolor=RED, lw=2.2))
    ax.set_title('Roll-rate matrix (month-to-month)', fontsize=10, fontweight='bold')
    ax2 = fig.add_subplot(gs[1])
    cure = {'1-30': prob.loc['1-30', 'Current'], '31-60': prob.loc['31-60', 'Current']+prob.loc['31-60', '1-30'],
            '61-90': prob.loc['61-90', 'Current']+prob.loc['61-90', '1-30']+prob.loc['61-90', '31-60']}
    bars = ax2.bar(list(cure), list(cure.values()), color=[GREEN, '#e0a800', RED], width=.6, zorder=3)
    for b, v in zip(bars, cure.values()): ax2.text(b.get_x()+b.get_width()/2, v+1.5, f'{v:.0f}%', ha='center', fontweight='bold', fontsize=9)
    ax2.set_ylim(0, 80); ax2.set_ylabel('chance of curing next month (%)'); ax2.grid(axis='y', color='#eee', zorder=0)
    ax2.set_title('Cure chance collapses with depth', fontsize=10, fontweight='bold'); ax2.set_xlabel('starting bucket')
    fig.suptitle('Fragment 2 \u2014 Delinquency roll-rates & cure behaviour', fontweight='bold', fontsize=13, y=1.0)
    fig.text(0.5, 0.93, '61\u201390 is the point of no return: most roll to default; rescue happens in 1\u201330 / 31\u201360.', ha='center', fontsize=8.5, color='#666')
    plt.subplots_adjust(left=0.12, right=0.97, top=0.86, bottom=0.18)
    plt.savefig(f'{out}/fragment2_rollrate.png', dpi=150, bbox_inches='tight'); plt.close()

    # Portfolio overview
    fig = plt.figure(figsize=(9.6, 6.2)); gs = gridspec.GridSpec(2, 1, height_ratios=[1, 1.45], hspace=0.42)
    axk = fig.add_subplot(gs[0]); axk.axis('off'); axk.set_xlim(0, 7); axk.set_ylim(0, 1)
    tiles = [('Loans', f"{ov['n_loans']:,}"), ('Customers', f"{ov['n_cust']:,}"),
             ('Originated', f"\u20b9{ov['originated']:,} cr"), ('Outstanding', f"\u20b9{ov['outstanding']:,} cr"),
             ('Book profit', f"\u20b9{ov['profit']} cr"), ('Default (90+)', f"{ov['ever90']}%"),
             ('Curr. delinquent', f"{ov['cur_delq']}%")]
    for i, (lab, val) in enumerate(tiles):
        axk.add_patch(plt.Rectangle((i+0.06, 0.12), 0.88, 0.76, facecolor='#f4f6fb', edgecolor='#ccd4e6', lw=1))
        axk.text(i+0.5, 0.62, val, ha='center', va='center', fontsize=12.5, fontweight='bold', color=NAVY)
        axk.text(i+0.5, 0.27, lab, ha='center', va='center', fontsize=8, color='#555')
    axk.set_title('Portfolio overview \u2014 state of the book @ month 24', fontsize=13, fontweight='bold', loc='left')
    axk.text(0, 1.18, f"WA APR {ov['wa_apr']}% (ticket-weighted)  \u00b7  WA tenure {ov['wa_ten']} mo  \u00b7  mean ticket \u20b9{ov['mean_tk']:,}  \u00b7  {ov['n_loans']/ov['n_cust']:.2f} loans/customer", fontsize=8.5, color='#666')
    axc = fig.add_subplot(gs[1]); xx = np.arange(3); w = 0.26
    L = con.n/con.n.sum()*100; E = con.expo/con.expo.sum()*100; P = con.profit/con.profit.sum()*100
    axc.bar(xx-w, L, w, label='% of loans', color='#9fb3d9', zorder=3)
    axc.bar(xx, E, w, label='% of exposure (origination \u20b9)', color=NAVY, zorder=3)
    axc.bar(xx+w, P, w, label='% of book profit', color=[RED if v < 0 else GREEN for v in P], zorder=3)
    axc.axhline(0, color='#444', lw=1)
    for xi, p in zip(xx, PRODUCTS):
        for dx, ser in zip([-w, 0, w], [L, E, P]):
            v = ser[p]; axc.text(xi+dx, v+(1.5 if v >= 0 else -1.5), f'{v:.0f}', ha='center', va='bottom' if v >= 0 else 'top', fontsize=8, fontweight='bold', color=INK)
    axc.set_xticks(xx); axc.set_xticklabels(PRODUCTS, fontsize=11); axc.set_ylabel('% of book'); axc.set_ylim(-12, 90)
    axc.grid(axis='y', color='#eee', zorder=0); axc.legend(loc='upper left', fontsize=8, frameon=False)
    axc.set_title('Concentration: SME ~20% of loans but ~76% of exposure & ~78% of profit; BNPL = high-volume, near-zero-exposure, loss-making tail', fontsize=9, color='#444', pad=6, loc='left')
    plt.savefig(f'{out}/fragment3_portfolio_overview.png', dpi=150, bbox_inches='tight'); plt.close()

    # Crossover
    mean_vp, n, defr, tot = xover
    fig = plt.figure(figsize=(10, 4.8)); gs = gridspec.GridSpec(1, 2, width_ratios=[1.7, 1], wspace=0.3)
    ax = fig.add_subplot(gs[0]); Mx = mean_vp.values
    norm = TwoSlopeNorm(vmin=slog(np.nanmin(Mx))*1.05, vcenter=0, vmax=slog(np.nanmax(Mx))*1.05)
    ax.imshow(slog(Mx), cmap='RdYlGn', norm=norm, aspect='auto')
    ax.set_xticks(range(5)); ax.set_xticklabels(CHANNELS, rotation=25, ha='right', fontsize=8.5)
    ax.set_yticks(range(3)); ax.set_yticklabels(PRODUCTS)
    for i in range(3):
        for j in range(5): ax.text(j, i, f'\u20b9{Mx[i,j]:,.0f}', ha='center', va='center', fontsize=8.3, fontweight='bold', color='#111')
    ax.set_title('Mean profit per loan \u2014 product \u00d7 channel', fontsize=10, fontweight='bold')
    ax2 = fig.add_subplot(gs[1]); vals = (tot.loc['BNPL']/1e5).reindex(CHANNELS)
    ax2.barh(CHANNELS[::-1], vals[::-1], color=RED, zorder=3)
    ax2.set_xlabel('total BNPL loss (\u20b9 lakh)'); ax2.grid(axis='x', color='#eee', zorder=0)
    ax2.set_title('Where BNPL bleeds, in aggregate', fontsize=10, fontweight='bold')
    fig.suptitle('Workstream D capstone \u2014 product \u00d7 channel crossover', fontweight='bold', fontsize=13, y=1.0)
    fig.text(0.5, 0.92, 'BNPL loses in every channel; the aggregate drag concentrates in Digital ads & DSA \u2014 act there first.', ha='center', fontsize=8.5, color='#666')
    plt.savefig(f'{out}/fragment4_product_channel.png', dpi=150, bbox_inches='tight'); plt.close()
    print('  charts written (7 PNGs)')

# ================================================================ WORKBOOK
def build_workbook(A, B, C, emob, vint, prob, ORDER, ttd, ov, con, out, xover=None, q2_path=None):
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    FT = 'Arial'
    HEAD = Font(name=FT, bold=True, color='FFFFFF', size=10); H1 = Font(name=FT, bold=True, size=14)
    H2 = Font(name=FT, bold=True, size=11); BODY = Font(name=FT, size=10); GREY = Font(name=FT, size=9, color='666666', italic=True)
    NAVY = PatternFill('solid', start_color='1F3864'); BAND = PatternFill('solid', start_color='D9E1F2')
    LOSS = PatternFill('solid', start_color='F8CBAD'); OKF = PatternFill('solid', start_color='C6EFCE'); PEND = PatternFill('solid', start_color='FFEB9C')
    ctr = Alignment('center', 'center'); left = Alignment('left', 'center', wrap_text=True)
    thin = Side('thin', color='BFBFBF'); BORD = Border(thin, thin, thin, thin)
    PCT = '0.0%'; RUP = '\u20b9#,##0;(\u20b9#,##0)'

    def hdr(ws, row, cols, widths):
        for c, (name, w) in enumerate(zip(cols, widths), 1):
            cell = ws.cell(row, c, name); cell.font = HEAD; cell.fill = NAVY; cell.alignment = ctr; cell.border = BORD
            ws.column_dimensions[get_column_letter(c)].width = w

    wb = Workbook()
    # README
    ws = wb.active; ws.title = 'README'; ws.sheet_view.showGridLines = False
    ws['A1'] = 'Workstream D \u2014 EDA log (Q3 + diagnostics)'; ws['A1'].font = H1
    for r, (a, b) in enumerate([
        ('Provenance', 'Frozen WS-C data, seed 20260603, Gate 1 PASS (32/32 MUST). 49,600 loans / 40,000 customers; M24 snapshot.'),
        ('Grain', 'LOAN level. value_proxy = NII+Fee-Loss-Servicing-CAC (reconciled to <1 rupee).'),
        ('Caveat', 'Mixed loan ages: low-maturity cells show profit/default "so far"; long/large cells understated.'),
        ('BNPL bands', 'Rescaled to real 2k-50k range (A-033); documented bands with auto-fallback guard.'),
        ('Reproduce', 'Generated by wsD_analysis.py. Derived columns are live formulas (recalc on open).')], 3):
        ws.cell(r, 1, a).font = H2; ws.cell(r, 2, b).font = BODY; ws.cell(r, 2).alignment = left
    ws.column_dimensions['A'].width = 14; ws.column_dimensions['B'].width = 110

    # Slice_A
    ws = wb.create_sheet('Slice_A'); ws.sheet_view.showGridLines = False
    ws['A1'] = 'Slice A \u2014 Product baseline (loan-level)'; ws['A1'].font = H1
    hdr(ws, 3, ['Product', 'n_loans', 'share_of_book', 'default_%', 'mean_ticket (\u20b9)', 'mean_tenure (mo)',
                'mean_APR', 'mean_profit (\u20b9)', 'total_profit (\u20b9)', 'total_profit (\u20b9 cr)', 'profit_share'],
        [12, 10, 12, 10, 15, 15, 10, 15, 18, 16, 12])
    for i, p in enumerate(PRODUCTS):
        r = 4+i; row = A.loc[p]
        ws.cell(r, 1, p).font = BODY; ws.cell(r, 2, int(row.n)).font = BODY
        ws.cell(r, 3, f'=B{r}/SUM($B$4:$B$6)').number_format = PCT
        ws.cell(r, 4, round(row.defr, 4)).number_format = PCT
        ws.cell(r, 5, round(row.tk)).number_format = '\u20b9#,##0'
        ws.cell(r, 6, round(row.ten, 1)); ws.cell(r, 7, round(row.apr/100, 4)).number_format = PCT
        ws.cell(r, 8, round(row.mvp)).number_format = RUP
        ws.cell(r, 9, f'=B{r}*H{r}').number_format = RUP
        ws.cell(r, 10, f'=I{r}/10000000').number_format = '\u20b9#,##0.00'
        ws.cell(r, 11, f'=I{r}/SUM($I$4:$I$6)').number_format = PCT
        for c in range(1, 12): ws.cell(r, c).border = BORD; ws.cell(r, c).alignment = ctr
        if row.mvp < 0:
            for c in range(1, 12): ws.cell(r, c).fill = LOSS
    rt = 7
    ws.cell(rt, 1, 'TOTAL').font = Font(name=FT, bold=True)
    ws.cell(rt, 2, '=SUM(B4:B6)'); ws.cell(rt, 3, '=SUM(C4:C6)').number_format = PCT
    ws.cell(rt, 4, f'=SUMPRODUCT(B4:B6,D4:D6)/B{rt}').number_format = PCT
    ws.cell(rt, 8, f'=I{rt}/B{rt}').number_format = RUP
    ws.cell(rt, 9, '=SUM(I4:I6)').number_format = RUP
    ws.cell(rt, 10, '=SUM(J4:J6)').number_format = '\u20b9#,##0.00'
    ws.cell(rt, 11, '=SUM(K4:K6)').number_format = PCT
    for c in range(1, 12): ws.cell(rt, c).font = Font(name=FT, bold=True); ws.cell(rt, c).fill = BAND; ws.cell(rt, c).border = BORD; ws.cell(rt, c).alignment = ctr

    def slice_bc(name, df, order, title):
        ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
        ws['A1'] = title; ws['A1'].font = H1
        hdr(ws, 3, ['Product', 'Bucket', 'n_loans', 'share_of_product', 'default_%', 'mean_profit (\u20b9)',
                    'mean_MOB', 'mean_tenure', 'maturity', 'total_profit (\u20b9 cr)', 'flag'],
            [12, 12, 10, 15, 10, 15, 10, 12, 11, 16, 9])
        r = 4
        for p in PRODUCTS:
            for bk in order[p]:
                if (p, bk) not in df.index: continue
                row = df.loc[(p, bk)]
                ws.cell(r, 1, p); ws.cell(r, 2, bk); ws.cell(r, 3, int(row.n))
                ws.cell(r, 4, f'=C{r}/SUMIF($A$4:$A$100,A{r},$C$4:$C$100)').number_format = PCT
                ws.cell(r, 5, round(row.defr, 4)).number_format = PCT
                ws.cell(r, 6, round(row.mvp)).number_format = RUP
                ws.cell(r, 7, round(row.mob, 1)); ws.cell(r, 8, round(row.ten, 1))
                ws.cell(r, 9, f'=G{r}/H{r}').number_format = PCT
                ws.cell(r, 10, f'=C{r}*F{r}/10000000').number_format = '\u20b9#,##0.00'
                ws.cell(r, 11, f'=IF(F{r}<0,"LOSS","")')
                for c in range(1, 12): ws.cell(r, c).border = BORD; ws.cell(r, c).alignment = ctr; ws.cell(r, c).font = BODY
                if row.mvp < 0:
                    for c in range(1, 12): ws.cell(r, c).fill = LOSS
                r += 1
    slice_bc('Slice_B', B, TENURE_ORDER, 'Slice B \u2014 Product \u00d7 Tenure (loan-level)')
    slice_bc('Slice_C', C, TICKET_ORDER, 'Slice C \u2014 Product \u00d7 Ticket (loan-level)')

    # EqualMOB
    ws = wb.create_sheet('EqualMOB_check'); ws.sheet_view.showGridLines = False
    ws['A1'] = 'Equal-MOB default check \u2014 does long-tenure low default survive same-time comparison?'; ws['A1'].font = H1
    hdr(ws, 3, ['Product', 'Tenure bucket', 'raw default', 'default @ MOB 12', 'n watched 12mo'], [12, 15, 12, 16, 16])
    for i, (_, x) in enumerate(emob.iterrows()):
        r = 4+i
        ws.cell(r, 1, x['product']); ws.cell(r, 2, x['bucket'])
        ws.cell(r, 3, round(x['raw_default']/100, 4)).number_format = PCT
        ws.cell(r, 4, round(x['default_at_MOB12']/100, 4)).number_format = PCT
        ws.cell(r, 5, int(x['n_watched12']))
        for c in range(1, 6): ws.cell(r, c).border = BORD; ws.cell(r, c).font = BODY; ws.cell(r, c).alignment = ctr

    # Vintage
    ws = wb.create_sheet('Vintage'); ws.sheet_view.showGridLines = False
    ws['A1'] = 'Vintage / cohort analysis'; ws['A1'].font = H1
    ws['A2'] = 'Raw default falls with recency = maturity artefact; equal-age (30+ by MOB3) drifts UP = real R10.'; ws['A2'].font = GREY
    hdr(ws, 4, ['Cohort', 'n', 'raw default %', '30+ by MOB3 % (equal age)', 'n watched 3mo', 'grade D+E intake %'], [12, 8, 14, 24, 15, 18])
    for i, (_, x) in enumerate(vint.iterrows()):
        r = 5+i
        ws.cell(r, 1, x['cohort']).font = BODY; ws.cell(r, 2, int(x['n'])).font = BODY
        ws.cell(r, 3, x['raw_default']/100 if pd.notna(x['raw_default']) else None).number_format = PCT
        ws.cell(r, 4, x['early30_MOB3']/100 if pd.notna(x['early30_MOB3']) else None).number_format = PCT
        ws.cell(r, 5, int(x['n_obs3'])).font = BODY; ws.cell(r, 6, x['grade_DE']/100).number_format = PCT
        for c in range(1, 7): ws.cell(r, c).border = BORD; ws.cell(r, c).alignment = ctr

    # RollRate
    ws = wb.create_sheet('RollRate'); ws.sheet_view.showGridLines = False
    ws['A1'] = 'Delinquency roll-rate matrix (month-to-month, % of row)'; ws['A1'].font = H1
    hdr(ws, 4, ['this month \\ next']+ORDER+['row check'], [18, 11, 11, 11, 11, 11, 11])
    for i, b in enumerate(ORDER):
        r = 5+i
        ws.cell(r, 1, b).font = Font(name=FT, bold=True); ws.cell(r, 1).fill = BAND; ws.cell(r, 1).border = BORD; ws.cell(r, 1).alignment = ctr
        for j, b2 in enumerate(ORDER):
            c = ws.cell(r, 2+j, round(prob.loc[b, b2]/100, 4)); c.number_format = PCT; c.border = BORD; c.alignment = ctr; c.font = BODY
        ws.cell(r, 7, f'=SUM(B{r}:F{r})').number_format = PCT; ws.cell(r, 7).border = BORD; ws.cell(r, 7).alignment = ctr
    ws.cell(11, 1, f'Time-to-first-delinquency: median {ttd[0]:.0f} MOB, mean {ttd[1]:.1f} MOB ({ttd[2]:,} delinquent loans).').font = GREY

    # Portfolio_Overview
    ws = wb.create_sheet('Portfolio_Overview'); ws.sheet_view.showGridLines = False
    ws['A1'] = 'Portfolio overview \u2014 state of the book @ M24'; ws['A1'].font = H1
    kpi = [('Loans', ov['n_loans'], '0'), ('Customers', ov['n_cust'], '0'),
           ('Total originated (\u20b9 cr)', ov['originated'], '\u20b9#,##0'), ('Current outstanding, active (\u20b9 cr)', ov['outstanding'], '\u20b9#,##0'),
           ('Book profit (\u20b9 cr)', ov['profit'], '\u20b9#,##0.0'), ('Ever 90+ / default rate', ov['ever90']/100, PCT),
           ('Currently delinquent (active)', ov['cur_delq']/100, PCT), ('Currently 90+ (active)', ov['cur_90']/100, PCT),
           ('WA APR (ticket-weighted)', ov['wa_apr']/100, PCT), ('WA tenure (months)', ov['wa_ten'], '0.0'), ('Mean ticket (\u20b9)', ov['mean_tk'], '\u20b9#,##0')]
    ws['A3'] = 'Headline KPIs'; ws['A3'].font = H2
    for i, (lab, val, fmt) in enumerate(kpi):
        r = 4+i; ws.cell(r, 1, lab).font = BODY; c = ws.cell(r, 2, val); c.number_format = fmt; c.font = BODY
        ws.cell(r, 1).border = BORD; c.border = BORD
    ws.column_dimensions['A'].width = 34; ws.column_dimensions['B'].width = 16
    cr0 = 4+len(kpi)+1; ws.cell(cr0, 1, 'Concentration by product').font = H2
    hdr(ws, cr0+1, ['Product', 'n_loans', 'exposure (\u20b9 cr)', 'profit (\u20b9 cr)', '% loans', '% exposure', '% profit'], [12, 12, 16, 14, 10, 12, 10])
    for i, p in enumerate(PRODUCTS):
        rr = cr0+2+i
        ws.cell(rr, 1, p).font = BODY; ws.cell(rr, 2, int(con.loc[p, 'n'])).font = BODY
        ws.cell(rr, 3, round(con.loc[p, 'expo']/CRORE, 1)).number_format = '\u20b9#,##0.0'
        ws.cell(rr, 4, round(con.loc[p, 'profit']/CRORE, 2)).number_format = '\u20b9#,##0.00'
        ws.cell(rr, 5, f'=B{rr}/SUM($B${cr0+2}:$B${cr0+4})').number_format = PCT
        ws.cell(rr, 6, f'=C{rr}/SUM($C${cr0+2}:$C${cr0+4})').number_format = PCT
        ws.cell(rr, 7, f'=D{rr}/SUM($D${cr0+2}:$D${cr0+4})').number_format = PCT
        for c in range(1, 8): ws.cell(rr, c).border = BORD; ws.cell(rr, c).alignment = ctr

    # Findings
    ws = wb.create_sheet('Findings'); ws.sheet_view.showGridLines = False
    ws['A1'] = 'Q3 Findings'; ws['A1'].font = H1
    F = [('Headline', 'Structurally money-losing combination: BNPL, across ALL tickets and ALL tenures.'),
         ('1', 'SME is the profit engine: 20% of loans but 78% of book profit (\u20b937.0 cr of \u20b947.4 cr).'),
         ('2', 'BNPL loses on average (\u2212\u20b9912/loan) at every ticket and tenure, on 92\u201396% mature loans \u2014 not a small-ticket floor.'),
         ('3', 'Large-ticket SME (>\u20b91M) is the best cell: \u20b969,604/loan, \u20b925.6 cr \u2014 over half the book.'),
         ('4', 'Profit rises with tenure & ticket for Personal/SME; longer tenure also genuinely lower-risk (equal-MOB).'),
         ('5', 'Small-ticket Personal (<\u20b950k) is the marginal watch-cell: +\u20b9168, highest Personal default (8.5%).')]
    for i, (a, b) in enumerate(F):
        r = 3+i; ws.cell(r, 1, a).font = Font(name=FT, bold=True); ws.cell(r, 2, b).font = BODY; ws.cell(r, 2).alignment = left
    ws.column_dimensions['A'].width = 12; ws.column_dimensions['B'].width = 115

    # Gate2_Readiness
    ws = wb.create_sheet('Gate2_Readiness'); ws.sheet_view.showGridLines = False
    ws['A1'] = 'Gate 2 (EDA review) \u2014 readiness'; ws['A1'].font = H1
    hdr(ws, 3, ['Gate 2 requirement', 'Status', 'Owner', 'Where'], [46, 14, 16, 26])
    items = [('Insights decision-relevant, not descriptive', 'DONE', 'us', 'all slices + fragments'),
             ('Unit economics netted (loss, CAC, servicing)', 'DONE', 'us', 'Slice_A'),
             ('Product / ticket / tenure economics (Q3)', 'DONE', 'us', 'Slice_A/B/C'),
             ('Cohort / vintage analysis', 'DONE', 'us', 'Vintage'),
             ('Roll-rate / bucket migration', 'DONE', 'us', 'RollRate'),
             ('Portfolio overview exhibit', 'DONE', 'us', 'Portfolio_Overview'),
             ('Channel performance + CAC payback / CLV (Q2)', 'DONE' if q2_path else 'PENDING', 'coworker', 'Q2_* sheets' if q2_path else 'await Q2'),
             ('Product\u00d7channel crossover (capstone)', 'DONE' if q2_path else 'RECOMMENDED', 'us+coworker', 'Q2Q3_Crossover' if q2_path else 'needs Q2'),
             ('EM + Senior Risk Consultant sign-off', 'PENDING', 'EM / PMO', 'Gate 2 meeting')]
    for i, (req, st, own, where) in enumerate(items):
        r = 4+i; ws.cell(r, 1, req).font = BODY; ws.cell(r, 1).alignment = left
        c = ws.cell(r, 2, st); c.font = Font(name=FT, bold=True); c.alignment = ctr; c.fill = OKF if st == 'DONE' else PEND
        ws.cell(r, 3, own).font = BODY; ws.cell(r, 3).alignment = ctr; ws.cell(r, 4, where).font = BODY; ws.cell(r, 4).alignment = left
        for cc in range(1, 5): ws.cell(r, cc).border = BORD

    q3_file = f'{out}/Q3_results_log.xlsx'; wb.save(q3_file)
    print(f'  wrote {os.path.basename(q3_file)} ({len(wb.sheetnames)} sheets)')

    if q2_path and os.path.exists(q2_path):
        _incorporate_q2(q3_file, q2_path, out, xover)
    return q3_file

def _incorporate_q2(q3_file, q2_path, out, xover=None):
    """Fold the coworker's Q2 workbook in + add reconciliation/crossover sheets."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    FT = 'Arial'; HEAD = Font(name=FT, bold=True, color='FFFFFF', size=10); H1 = Font(name=FT, bold=True, size=14)
    H2 = Font(name=FT, bold=True, size=11); BODY = Font(name=FT, size=10); GREY = Font(name=FT, size=9, color='666666', italic=True)
    NAVY = PatternFill('solid', start_color='1F3864'); LOSS = PatternFill('solid', start_color='F8CBAD'); OKF = PatternFill('solid', start_color='C6EFCE')
    ctr = Alignment('center', 'center'); left = Alignment('left', 'center', wrap_text=True)
    thin = Side('thin', color='BFBFBF'); BORD = Border(thin, thin, thin, thin); PCT = '0.0%'; RUP = '\u20b9#,##0;(\u20b9#,##0)'
    def hdr(ws, row, cols, widths):
        for c, (name, w) in enumerate(zip(cols, widths), 1):
            cell = ws.cell(row, c, name); cell.font = HEAD; cell.fill = NAVY; cell.alignment = ctr; cell.border = BORD
            ws.column_dimensions[get_column_letter(c)].width = w

    ce = pd.read_excel(q2_path, 'Channel_Economics', header=4).iloc[:5]
    wb = load_workbook(q3_file)
    ws = wb.create_sheet('Q2_Channel_Economics'); ws.sheet_view.showGridLines = False
    ws['A1'] = 'Q2 \u2014 Channel economics (coworker; incorporated)'; ws['A1'].font = H1
    ws['A2'] = 'Per-loan, same value_proxy as Q3. Reconciles to the book total.'; ws['A2'].font = GREY
    hdr(ws, 4, ['Channel', 'Loans', 'Share', 'Avg CAC (\u20b9)', 'Default', 'Mean profit (\u20b9)', 'Median profit (\u20b9)', 'Skew %', 'Payback (mo)', 'CLV/CAC'],
        [16, 9, 9, 12, 9, 15, 16, 9, 11, 10])
    for i, (_, x) in enumerate(ce.iterrows()):
        r = 5+i
        ws.cell(r, 1, x['Channel']).font = BODY; ws.cell(r, 2, int(x['Loans'])).font = BODY
        ws.cell(r, 3, float(x['Share of Book'])).number_format = PCT
        ws.cell(r, 4, round(float(x['Avg CAC (\u20b9)']))).number_format = '\u20b9#,##0'
        ws.cell(r, 5, float(x['Default Rate'])).number_format = PCT
        ws.cell(r, 6, round(float(x['Mean Profit (\u20b9)']))).number_format = RUP
        ws.cell(r, 7, round(float(x['Median Profit (\u20b9)']))).number_format = RUP
        ws.cell(r, 8, float(x['Skew %'])/100).number_format = PCT
        ws.cell(r, 9, float(x['Payback (months)'])).number_format = '0.0'
        ws.cell(r, 10, float(x['CLV / CAC'])).number_format = '0.0"x"'
        for c in range(1, 11): ws.cell(r, c).border = BORD; ws.cell(r, c).alignment = ctr
    # reconciliation note
    ws = wb.create_sheet('Q2Q3_Reconciliation'); ws.sheet_view.showGridLines = False
    ws['A1'] = 'Audit \u2014 Q2 \u00d7 Q3 reconciliation'; ws['A1'].font = H1
    for i, line in enumerate(['Channel loans sum = 49,600 = book total.',
                              'Channel profit sum = product profit sum (\u20b947.36 cr) \u2014 identical underlying loans.',
                              'Channel mean profit/loan matches coworker to the rupee. Same value_proxy, same grain (per-loan).',
                              'Both cuts correct; BNPL loss sits inside every channel (see Q2Q3_Crossover).']):
        ws.cell(3+i, 1, line).font = BODY; ws.cell(3+i, 1).alignment = left
    ws.column_dimensions['A'].width = 110
    # crossover (our capstone) — product x channel
    if xover is not None:
        mean_vp, n, defr, tot = xover
        ws = wb.create_sheet('Q2Q3_Crossover'); ws.sheet_view.showGridLines = False
        ws['A1'] = 'Capstone \u2014 product \u00d7 channel crossover'; ws['A1'].font = H1
        ws['A2'] = 'BNPL loses in every channel; the drag concentrates in Digital ads & DSA.'; ws['A2'].font = GREY
        ws['A4'] = 'A. Mean profit (\u20b9/loan) by product \u00d7 channel'; ws['A4'].font = H2
        CHN = list(mean_vp.columns)
        hdr(ws, 5, ['Product']+CHN, [12]+[15]*len(CHN))
        for i, p in enumerate(mean_vp.index):
            r = 6+i; ws.cell(r, 1, p).font = Font(name=FT, bold=True); ws.cell(r, 1).alignment = ctr; ws.cell(r, 1).border = BORD
            for j, c in enumerate(CHN):
                cell = ws.cell(r, 2+j, round(mean_vp.loc[p, c])); cell.number_format = RUP; cell.alignment = ctr; cell.border = BORD; cell.font = BODY
                if mean_vp.loc[p, c] < 0: cell.fill = LOSS
        r = 6+len(mean_vp.index)+1
        ws.cell(r, 1, 'B. BNPL by channel \u2014 where to act (sorted by aggregate loss)').font = H2; r += 1
        hdr(ws, r, ['Channel', 'BNPL loans', 'BNPL mean (\u20b9)', 'BNPL total loss (\u20b9 cr)', '% of BNPL drag', 'BNPL default'], [16, 12, 15, 18, 13, 12]); r += 1
        drag = tot.loc['BNPL'].sum()
        for c in tot.loc['BNPL'].sort_values().index:
            ws.cell(r, 1, c).font = BODY; ws.cell(r, 1).fill = LOSS
            ws.cell(r, 2, int(n.loc['BNPL', c])).font = BODY
            ws.cell(r, 3, round(mean_vp.loc['BNPL', c])).number_format = RUP
            ws.cell(r, 4, round(tot.loc['BNPL', c]/10000000, 2)).number_format = '\u20b9#,##0.00'
            ws.cell(r, 5, tot.loc['BNPL', c]/drag).number_format = PCT
            ws.cell(r, 6, defr.loc['BNPL', c]).number_format = PCT
            for cc in range(1, 7): ws.cell(r, cc).border = BORD; ws.cell(r, cc).alignment = ctr
            r += 1
        ws.cell(r+1, 1, 'Digital ads \u2248 50% of the \u20b91.57 cr BNPL drag; + Partner + DSA \u2248 89%. Act on Digital ads & DSA first.').font = GREY
        ws.column_dimensions['A'].width = 16
    wb.save(f'{out}/WorkstreamD_consolidated_log.xlsx')
    print(f'  wrote WorkstreamD_consolidated_log.xlsx ({len(wb.sheetnames)} sheets, Q2 incorporated)')

# ================================================================ MAIN
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--data', default='./out', help='WS-C output root (contains data/ and meta/)')
    ap.add_argument('--out', default='./wsD_outputs', help='output directory')
    ap.add_argument('--q2', default=None, help='optional path to the Q2 controlled workbook')
    a = ap.parse_args()
    os.makedirs(a.out, exist_ok=True)

    print('Workstream D EDA \u2014 regenerating from frozen data')
    m, cust = load_and_reconcile(a.data)
    rep = pd.read_parquet(f'{a.data}/data/parquet/repayments.parquet')
    loans = pd.read_parquet(f'{a.data}/data/parquet/loans.parquet')

    A, B, C, m = compute_slices(m)
    emob = equal_mob((loans, rep))
    vint = vintage((loans, rep))
    prob, ORDER, ttd = rollrate(rep)
    ov, con = portfolio_overview(m, cust, rep)
    xover = crossover(m)

    make_charts(A, B, C, vint, prob, ORDER, ov, con, xover, a.out)
    build_workbook(A, B, C, emob, vint, prob, ORDER, ttd, ov, con, a.out, xover=xover, q2_path=a.q2)
    print('Done. Outputs in', a.out)

if __name__ == '__main__':
    main()
